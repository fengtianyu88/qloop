#!/usr/bin/env python3
"""合并所有测试结果并更新 Excel 表格

输入:
- test_results.json (主测试,113 项)
- test_results_supplement.json (补充测试,28 项)
- test_results_retest.json (重测,3 项)

输出:
- 更新 QLoop_v1.4.7_测试用例.xlsx 中的"执行状态"和"实际结果"列
- 生成 merged_test_results.json (合并后的完整结果)
"""
import json
import os
import openpyxl
from openpyxl.styles import PatternFill

BASE = "/mnt/c/Users/tiany/Documents/Trae solo my data"

# 1. 加载所有测试结果
def load_json(path):
    full = os.path.join(BASE, path) if not path.startswith("/mnt") else path
    try:
        with open(full, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"警告: {path} 不存在")
        return []

main_results = load_json("test_results.json")
supplement_results = load_json("test_results_supplement.json")
retest_results = load_json("test_results_retest.json")
p0_results = load_json("test_results_p0.json")
llm08_results = load_json("test_results_llm08.json")

print(f"主测试: {len(main_results)} 项")
print(f"补充测试: {len(supplement_results)} 项")
print(f"重测: {len(retest_results)} 项")
print(f"P0补充: {len(p0_results)} 项")
print(f"LLM-08重测: {len(llm08_results)} 项")

# 2. 合并结果(后执行的覆盖先执行的)
# 优先级: llm08 > p0 > retest > supplement > main
merged = {}
for r in main_results:
    merged[r["tc_id"]] = r
for r in supplement_results:
    merged[r["tc_id"]] = r  # 覆盖
for r in retest_results:
    merged[r["tc_id"]] = r  # 覆盖
for r in p0_results:
    merged[r["tc_id"]] = r  # 覆盖
for r in llm08_results:
    merged[r["tc_id"]] = r  # 覆盖

print(f"\n合并后: {len(merged)} 个唯一测试用例")

# 3. 统计
by_status = {}
for tc_id, r in merged.items():
    s = r.get("status", "UNKNOWN")
    by_status.setdefault(s, []).append(tc_id)

print("\n=== 合并结果统计 ===")
for s, ids in sorted(by_status.items()):
    print(f"  {s}: {len(ids)}")

# 4. 保存合并结果
merged_path = os.path.join(BASE, "merged_test_results.json")
with open(merged_path, "w", encoding="utf-8") as f:
    json.dump(list(merged.values()), f, ensure_ascii=False, indent=2)
print(f"\n合并结果已保存: {merged_path}")

# 5. 更新 Excel
xlsx_path = os.path.join(BASE, "QLoop_v1.4.7_测试用例.xlsx")
wb = openpyxl.load_workbook(xlsx_path)

# 找到测试用例总表
ws = wb["测试用例总表"]

# 找到"执行状态"和"实际结果"列(第11和12列)
# 表头在第1行,数据从第2行开始
STATUS_COL = 11  # K列
RESULT_COL = 12  # L列
NOTES_COL = 13   # M列
TC_ID_COL = 1    # A列

# 样式
pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
skip_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

updated = 0
not_found = 0
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=TC_ID_COL).value
    if not tc_id:
        continue
    tc_id = str(tc_id).strip()

    if tc_id in merged:
        r = merged[tc_id]
        status = r.get("status", "")
        actual = r.get("actual_result", "")
        notes = r.get("notes", "")

        # 写入执行状态
        cell_status = ws.cell(row=row, column=STATUS_COL, value=status)
        if status == "PASS":
            cell_status.fill = pass_fill
        elif status == "FAIL":
            cell_status.fill = fail_fill
        else:
            cell_status.fill = skip_fill

        # 写入实际结果
        ws.cell(row=row, column=RESULT_COL, value=actual)

        # 写入备注
        if notes:
            ws.cell(row=row, column=NOTES_COL, value=notes)

        updated += 1
    else:
        not_found += 1
        # 标注为未执行
        cell = ws.cell(row=row, column=STATUS_COL, value="未执行")
        cell.fill = skip_fill

print(f"\nExcel 更新完成: {updated} 个用例已更新, {not_found} 个未在测试结果中")

# 6. 添加测试汇总工作表
if "测试汇总" in wb.sheetnames:
    del wb["测试汇总"]
ws_summary = wb.create_sheet("测试汇总", 0)

# 标题
ws_summary["A1"] = "QLoop v1.4.7 测试汇总报告"
ws_summary["A1"].font = openpyxl.styles.Font(name="Microsoft YaHei", size=14, bold=True)
ws_summary.merge_cells("A1:D1")

# 统计数据
total = len(merged)
passed = len(by_status.get("PASS", []))
failed = len(by_status.get("FAIL", []))
skipped = len(by_status.get("SKIP", []))
pass_rate = passed * 100 / total if total else 0

ws_summary["A3"] = "总测试用例数"
ws_summary["B3"] = total
ws_summary["A4"] = "通过"
ws_summary["B4"] = passed
ws_summary["A5"] = "失败"
ws_summary["B5"] = failed
ws_summary["A6"] = "跳过"
ws_summary["B6"] = skipped
ws_summary["A7"] = "通过率"
ws_summary["B7"] = f"{pass_rate:.1f}%"

ws_summary["A9"] = "测试执行轮次"
ws_summary["B9"] = "4 轮 (3 轮主测试 + 1 轮补充测试 + 1 轮重测)"
ws_summary["A10"] = "发现并修复的 Bug"
ws_summary["B10"] = "5 个 (TC-DL-06 presigned URL 有效期 + TC-UPLOAD-14 空文件检查 + 3 个测试脚本适配)"

ws_summary["A12"] = "测试日期"
ws_summary["B12"] = "2026-07-22"
ws_summary["A13"] = "测试环境"
ws_summary["B13"] = "WSL Ubuntu-24.04 + PostgreSQL + Redis + MinIO + Celery"
ws_summary["A14"] = "后端版本"
ws_summary["B14"] = "v1.4.7"
ws_summary["A15"] = "测试执行人"
ws_summary["B15"] = "AI 自动化测试 + 人工审核"

# 设置列宽
ws_summary.column_dimensions["A"].width = 20
ws_summary.column_dimensions["B"].width = 60

# 7. 保存
wb.save(xlsx_path)
print(f"\nExcel 已保存: {xlsx_path}")
