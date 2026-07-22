#!/usr/bin/env python3
"""检查哪些 Excel 中的测试用例没有被执行,并标注原因"""
import json
import os
import openpyxl

BASE = "/mnt/c/Users/tiany/Documents/Trae solo my data"

# 加载合并后的测试结果
with open(os.path.join(BASE, "merged_test_results.json"), "r", encoding="utf-8") as f:
    executed = {r["tc_id"] for r in json.load(f)}

# 加载 Excel
wb = openpyxl.load_workbook(os.path.join(BASE, "QLoop_v1.4.7_测试用例.xlsx"))
ws = wb["测试用例总表"]

unexecuted = []
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row=row, column=1).value
    if not tc_id:
        continue
    tc_id = str(tc_id).strip()
    if tc_id not in executed:
        module = ws.cell(row=row, column=2).value or ""
        title = ws.cell(row=row, column=4).value or ""
        priority = ws.cell(row=row, column=5).value or ""
        unexecuted.append((tc_id, module, priority, title))

print(f"未执行的测试用例: {len(unexecuted)} 个")
print(f"\n{'TC ID':20} {'模块':15} {'优先级':5} {'标题'}")
print("-" * 80)
for tc_id, module, priority, title in unexecuted:
    print(f"{tc_id:20} {module:15} {priority:5} {title[:40]}")

# 按优先级统计
by_priority = {}
for _, _, priority, _ in unexecuted:
    by_priority.setdefault(priority, []).append(_)
print(f"\n按优先级统计:")
for p, ids in sorted(by_priority.items()):
    print(f"  {p}: {len(ids)} 个")
