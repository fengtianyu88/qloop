"""Batch import API - templates and imports for projects/users/orgs."""

import io
import uuid
from typing import List, Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from passlib.context import CryptContext
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user, require_roles
from app.models.organization import AdminScope, OrgUnit
from app.models.project import Project
from app.models.user import SystemRole, User
from app.schemas.organization import OrgUnitCreate
from app.schemas.project import ProjectCreate
from app.schemas.user import UserCreate
from app.services.audit_service import create_audit_log
from app.services.org_service import create_org_unit
from app.services.project_service import create_project
from app.services.user_service import create_user


router = APIRouter(prefix="/api/import", tags=["import"])

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# ============================================================
# Templates (download)
# ============================================================

def _make_projects_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.append(["项目名称", "描述"])
    ws.append(["BMS 算法 V2", "电池管理系统算法 V2 版本"])
    ws.append(["SOX 合规工具", "SOX 审计辅助工具"])
    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_users_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(["用户名", "邮箱", "姓名", "密码", "系统角色", "部门", "科室"])
    ws.append(["dev_zhang", "zhang@example.com", "张三", "Pass1234", "developer", "研发部", "算法科"])
    ws.append(["tester_li", "li@example.com", "李四", "Pass1234", "developer", "测试部", "功能测试科"])
    ws.append(["admin_wang", "wang@example.com", "王五", "Pass1234", "admin", "运维部", ""])
    for col in "ABCDEFG":
        ws.column_dimensions[col].width = 20
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_orgs_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OrgUnits"
    ws.append(["名称", "类型", "上级组织名称(可空)", "描述"])
    ws.append(["总经办", "department", "", "公司最高决策层"])
    ws.append(["研发中心", "division", "总经办", "负责产品研发"])
    ws.append(["算法部", "group", "研发中心", "BMS/SOX 算法开发"])
    for col in "ABCD":
        ws.column_dimensions[col].width = 25
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/projects/template")
async def download_projects_template(
    current_user: User = Depends(get_current_user),
):
    """Download Excel template for project import."""
    data = _make_projects_template()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=projects_template.xlsx"},
    )


@router.get("/users/template")
async def download_users_template(
    current_user: User = Depends(
        require_roles(SystemRole.ADMIN, SystemRole.SUPER_ADMIN)
    ),
):
    """Download Excel template for user import (admin only)."""
    data = _make_users_template()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"},
    )


@router.get("/organizations/template")
async def download_orgs_template(
    current_user: User = Depends(
        require_roles(SystemRole.ADMIN, SystemRole.SUPER_ADMIN)
    ),
):
    """Download Excel template for organization import (admin only)."""
    data = _make_orgs_template()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=orgs_template.xlsx"},
    )


# ============================================================
# Batch import
# ============================================================


class ImportResult:
    def __init__(self):
        self.success: int = 0
        self.failed: int = 0
        self.errors: List[str] = []

    def add_success(self):
        self.success += 1

    def add_error(self, msg: str):
        self.failed += 1
        self.errors.append(msg)


@router.post("/projects")
async def import_projects(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Batch import projects from Excel."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = ImportResult()
    for idx, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        description = str(row[1]).strip() if len(row) > 1 and row[1] else None
        try:
            await create_project(
                db=db,
                project_create=ProjectCreate(name=name, description=description),
                pm_user_id=current_user.id,
            )
            result.add_success()
        except Exception as exc:
            result.add_error(f"Row {idx} ({name}): {exc}")

    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="import_projects",
        resource_type="project",
        resource_id="batch",
        details={"success": result.success, "failed": result.failed},
    )

    return {
        "success": result.success,
        "failed": result.failed,
        "errors": result.errors[:20],  # Limit error list
    }


@router.post("/users")
async def import_users(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_roles(SystemRole.ADMIN, SystemRole.SUPER_ADMIN)
    ),
):
    """Batch import users from Excel (admin only)."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = ImportResult()
    for idx, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            username = str(row[0]).strip()
            email = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            full_name = str(row[2]).strip() if len(row) > 2 and row[2] else username
            password = str(row[3]).strip() if len(row) > 3 and row[3] else "Pass1234"
            role_str = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "developer"
            department = str(row[5]).strip() if len(row) > 5 and row[5] else None
            section = str(row[6]).strip() if len(row) > 6 and row[6] else None

            # Validate role
            try:
                system_role = SystemRole(role_str)
            except ValueError:
                system_role = SystemRole.DEVELOPER

            await create_user(
                db=db,
                user_create=UserCreate(
                    username=username,
                    email=email,
                    full_name=full_name,
                    password=password,
                    system_role=system_role,
                    department=department,
                    section=section,
                ),
            )
            result.add_success()
        except Exception as exc:
            result.add_error(f"Row {idx} ({row[0] if row else '?'}): {exc}")

    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="import_users",
        resource_type="user",
        resource_id="batch",
        details={"success": result.success, "failed": result.failed},
    )

    return {
        "success": result.success,
        "failed": result.failed,
        "errors": result.errors[:20],
    }


@router.post("/organizations")
async def import_orgs(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(
        require_roles(SystemRole.ADMIN, SystemRole.SUPER_ADMIN)
    ),
):
    """Batch import organization units from Excel (admin only).

    Two-pass strategy:
    1. Create all units (parent_id=None)
    2. Update parent_id based on parent name
    """
    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be .xlsx or .xls")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel: {exc}")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    result = ImportResult()
    name_to_id = {}

    # Pass 1: create all units (no parent)
    for idx, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue
        try:
            name = str(row[0]).strip()
            type_str = str(row[1]).strip().lower() if len(row) > 1 and row[1] else "department"
            # v1.5.2: org_type 改为字符串,由 create_org_unit 校验是否存在
            org_type = type_str
            description = str(row[3]).strip() if len(row) > 3 and row[3] else None

            unit = await create_org_unit(
                db=db,
                org_create=OrgUnitCreate(
                    name=name,
                    org_type=org_type,
                    parent_id=None,
                    description=description,
                ),
            )
            name_to_id[name] = unit.id
            result.add_success()
        except Exception as exc:
            result.add_error(f"Row {idx} ({row[0] if row else '?'}): {exc}")

    # Pass 2: update parent_id based on parent name
    for idx, row in enumerate(rows, start=2):
        if not row or not row[0] or len(row) < 3 or not row[2]:
            continue
        try:
            name = str(row[0]).strip()
            parent_name = str(row[2]).strip()
            if parent_name and parent_name in name_to_id:
                parent_id = name_to_id[parent_name]
                # Update the org unit's parent_id
                from sqlalchemy import update
                stmt = (
                    update(OrgUnit)
                    .where(OrgUnit.name == name)
                    .where(OrgUnit.parent_id.is_(None))
                    .values(parent_id=parent_id)
                )
                await db.execute(stmt)
                await db.commit()
        except Exception as exc:
            result.add_error(f"Parent-link row {idx} ({row[0]}): {exc}")

    await create_audit_log(
        db=db,
        user_id=current_user.id,
        action="import_orgs",
        resource_type="org_unit",
        resource_id="batch",
        details={"success": result.success, "failed": result.failed},
    )

    return {
        "success": result.success,
        "failed": result.failed,
        "errors": result.errors[:20],
    }
